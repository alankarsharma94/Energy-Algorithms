# -*- coding: utf-8 -*-
"""
Created on Fri Jul 10 09:14:51 2020

@author: Alankar Sharma
"""
import time
import pandas as pd
import numpy as np
from openpyxl import load_workbook

""" 
Hedge Strategy Simulator 
Volume procurements under 3 types of protocols evaluated
Programmatic protocol - Irrespective of daily prices, hedge farthest months with stepwise hedge targets
Objectives protocol - In the near/middle timeframe, if prices fall under a limit or go higher than a limit,
                      hedge to a certain %age volume. 
                      In the very near term, if under 98% hedged, force hedges to 98%
"""
t = time.time()

strat_number = 6 # Different strategies evaluate different mixes of protocols
iterations = 2 #Iterations of price paths evaluated- Max = 200

for strat in range(1,strat_number+1):
    print("processing strategy "+str(strat))
    settle_prices = pd.read_excel("Inputs.xlsx", sheet_name='SettleMatrix') #End of month settled prices (for assessing payoff)
    daily_prices = pd.read_excel("Inputs.xlsx", sheet_name='Input Prices') #Daily DA simulated price input (for tracking hedge cost)
    input_volumes = pd.read_excel("Inputs.xlsx", sheet_name='Input Volumes') #Target full requirement volumes
    parameters = pd.read_excel("Inputs.xlsx", sheet_name='Parameters ' +str(strat)) #Price levels to use and %of vol to hedge
    parameters_backup = parameters.iloc[:,:]
    hedge_template = pd.read_excel("Hedge_Template.xlsx", sheet_name='Template') #Dataframe to configure processing setup
    existing_trades = pd.read_excel("Inputs.xlsx", sheet_name='Existing Trades') #Existing hedges to net against full req
    
    #Process Programmatic
    print("starting programmatic")
    month_diff = hedge_template.iloc[:,:]
    
    month_diff_col = month_diff.iloc[0,4:len(hedge_template.columns)]
    month_diff_col = month_diff_col.to_frame()
    month_diff_colT = month_diff_col.T
    month_diff_row = month_diff.iloc[2:len(hedge_template),1]
    month_diff_row = month_diff_row.to_frame()
    
    month_diff_cols = month_diff_colT.append([month_diff_colT]*(len(month_diff_row)-1),ignore_index=True)
    month_diff_rows = pd.concat([month_diff_row] * len(month_diff_col), axis=1)
    
    month_diff_cols.columns = range(month_diff_cols.shape[1])
    df_cols_name = month_diff_cols.columns
    month_diff_rows.columns = df_cols_name
    month_diff_rows = month_diff_rows.reset_index(drop = True)
    
    #Fix month_diff
    month_diff_temp = month_diff.iloc[2:,4:]
    month_diff_temp = month_diff_temp.reset_index(drop=True)
    month_diff_temp.columns = range(month_diff_temp.shape[1])
    month_diff_monthout = month_diff.iloc[2:,1]
    month_diff_monthout = month_diff_monthout.reset_index(drop=True)
    month_diff_monthout_temp = month_diff_temp.iloc[:,:]
    month_diff_monthout_temp = month_diff_monthout_temp.reset_index(drop=True)
    month_diff_monthfwd = month_diff.iloc[0,4:]
    month_diff_monthfwd = month_diff_monthfwd.reset_index(drop=True)
    month_diff_monthfwd_temp = month_diff_temp.iloc[:,:]
    month_diff_monthfwd_temp=month_diff_monthfwd_temp.reset_index(drop=True)
    
    for i in range(0, len(month_diff_temp)):
        month_diff_monthfwd_temp.iloc[i,:] = month_diff_monthfwd[:]
    for i in range(0, len(month_diff_temp.columns)):
        month_diff_monthout_temp.iloc[:,i] = month_diff_monthout[:]
    
    month_diff_temp = month_diff_monthfwd_temp.iloc[:,:] - month_diff_monthout_temp.iloc[:,:]
    month_diff = []
    month_diff = month_diff_temp.iloc[:,:]
    
    #Compute Existing Hedges
    budget_volumes_table = pd.DataFrame(0, index=np.arange(len(month_diff)), columns=month_diff.columns)
    existing_hedges_table = pd.DataFrame(0, index=np.arange(len(month_diff)), columns=month_diff.columns)
    hedge_ratio_existing = pd.DataFrame(0, index=np.arange(len(month_diff)), columns=month_diff.columns)
    hedge_vols_existing = pd.DataFrame(0, index=np.arange(len(month_diff)), columns=month_diff.columns)
    
    for i in range(0, len(month_diff)):
        budget_volumes_table.iloc[i,:] = input_volumes.iloc[:,1]
        existing_hedges_table.iloc[i,:] = existing_trades.iloc[:,1]
        hedge_ratio_existing.iloc[i,:] = existing_trades.iloc[:,1] / input_volumes.iloc[:,1]
        hedge_vols_existing.iloc[i,:] = existing_trades.iloc[:,1]
    
    #Consolidate budget volumes
    budget_compressed = pd.DataFrame(np.zeros((36, 36)))
    for i in range(0,36):
        start_day = 21*i
        end_day = 21*(i+1)-1
        budget_compressed.iloc[i,:] = budget_volumes_table.iloc[end_day,:]
    existing_compressed = pd.DataFrame(np.zeros((36, 36)))
    for i in range(0,36):
        start_day = 21*i
        end_day = 21*(i+1)-1
        existing_compressed.iloc[i,:] = existing_hedges_table.iloc[end_day,:]
    
    #Programmatic Hedging
    #Isolate prog hede targets
    month_diff_binary = month_diff.iloc[:,:]
    month_diff_binary = month_diff.mask(month_diff < 0, 0)
    
    prog_input = parameters.iloc[:,:]
    prog_last_month_index = (prog_input.Category.values == 'P').argmax()
    
    for i in range(0,prog_last_month_index):
        prog_input.iloc[i] = prog_input.iloc[prog_last_month_index]
        
    for i in range(0, len(prog_input)):
        prog_input.iloc[i,0] = i+1
    
    #Put on hedges
    hedge_ratio_incremental = pd.DataFrame(0, index=np.arange(len(hedge_ratio_existing)), columns=hedge_ratio_existing.columns)
    hedge_incremental_vol = pd.DataFrame(0, index=np.arange(len(hedge_ratio_existing)), columns=hedge_ratio_existing.columns)
    hedge_track = pd.DataFrame(0, index=np.arange(len(hedge_ratio_existing)), columns=hedge_ratio_existing.columns)
    
    month_diff_binary_unique = month_diff_binary.drop_duplicates()
    month_indexes = month_diff_binary_unique.index.values
    
    prog_input_table = pd.DataFrame(0, index=np.arange(len(hedge_ratio_existing)), columns=hedge_ratio_existing.columns)
    for i in range(0,len(hedge_ratio_existing)):
        prog_input_table.iloc[i,:] = prog_input.iloc[:,2]
    
    prog_final_ratios = pd.DataFrame(0, index=np.arange(len(hedge_ratio_existing)), columns=hedge_ratio_existing.columns)
    
    for k in range(0,len(prog_input)):
        days_start = 21*k
        days_end = 21* (k+1)
        prog_month_target = prog_input_table.iloc[days_start:days_end,:]
        prog_month_target = prog_month_target.shift(periods = k, axis = "columns", fill_value = 0)
        prog_final_ratios.iloc[days_start:days_end,:] = prog_month_target
        prog_month_target = []
    
    for i in range(0,len(hedge_ratio_existing)):
        hedge_ratio_incremental.iloc[i,:] = (prog_final_ratios.iloc[i,:]).subtract(hedge_ratio_existing.iloc[i,:])
        hedge_incremental_vol.iloc[i,:] = hedge_ratio_incremental.iloc[i,:].multiply(input_volumes.iloc[:,1])
        hedge_track.iloc[i,:] = hedge_incremental_vol.iloc[i,:].add(hedge_vols_existing.iloc[i,:])
    
    hedge_ratio_incremental[hedge_ratio_incremental < 0] = 0
    hedge_incremental_vol[hedge_incremental_vol < 0] = 0
    
    #Hedge_Cumulative_Vols creation
    hedge_cumulative_vol2 = hedge_incremental_vol.copy(deep = True)
    hedge_cumulative_vol = hedge_cumulative_vol2.copy(deep = True)
    for x in range(0,len(hedge_cumulative_vol2.columns)):
        series_inc = hedge_cumulative_vol2.iloc[:,x]
        series_inc_list = series_inc.tolist()
        #Find last non zero
        if np.count_nonzero(series_inc) == 0:
            series_inc = series_inc
        else:
            last_index = np.max(np.nonzero(series_inc_list))
            series_inc[last_index:] = series_inc_list[last_index]
        hedge_cumulative_vol.iloc[:,x] = series_inc
    
    hedge_track = hedge_cumulative_vol.add(hedge_vols_existing)
    ratio_track = pd.DataFrame(0, index=np.arange(len(hedge_ratio_existing)), columns=hedge_ratio_existing.columns)
    ratio_track.iloc[:,:] = hedge_track.iloc[:,:]/budget_volumes_table.iloc[:,:]
    
    #Programmatic Costs
    prog_prices_zeroes = pd.DataFrame(0, index=np.arange(len(hedge_ratio_existing)), columns=hedge_ratio_existing.columns)
    prog_hedges_zeroes = pd.DataFrame(0, index=np.arange(len(hedge_ratio_existing)), columns=hedge_ratio_existing.columns)
    daily_price_alliters_prog = daily_prices.iloc[1:,4:]
    daily_price_alliters_prog = daily_price_alliters_prog.reset_index(drop = True)
    iter = iterations
    prog_cost_sum = pd.DataFrame(np.zeros((iter, 36)))
    prog_vol_sum= pd.DataFrame(np.zeros((iter, 36)))
    
    for i in range(0,iter):
        iters_start = 756*i
        iters_end = 756*(i+1)
        daily_prices_iter = daily_price_alliters_prog.iloc[iters_start:iters_end,:]
        daily_prices_iter.columns = prog_prices_zeroes.columns
        #Loop pver days
        for j in range(0,35):
            day_start = 21*j
            day_end = 21* (j+1)
            prog_prices_zeroes.iloc[day_start,:] = daily_prices_iter.iloc[day_start,:]
            prog_inc_vol_fwd = hedge_incremental_vol.iloc[day_end,:]
            prog_inc_vol_bck = hedge_incremental_vol.iloc[day_start,:]
            prog_hedges_zeroes.iloc[day_end,:] = prog_inc_vol_fwd.subtract(prog_inc_vol_bck)
        prog_hedges_zeroes.iloc[0,:] =  hedge_incremental_vol.iloc[0,:]
        prog_hedges_zeroes[prog_hedges_zeroes<0] = 0
        prog_cost = prog_prices_zeroes.multiply(prog_hedges_zeroes)
        #Create Monthly Cost
        for k in range(0,36):
            prog_cost_monthly = prog_cost.iloc[:,k]
            prog_cost_sum.iloc[i,k] = prog_cost_monthly.values.sum()
            prog_vol_monthly = prog_hedges_zeroes.iloc[:,k]
            prog_vol_sum.iloc[i,k] = prog_vol_monthly.values.sum()
    
    #Compressed Programmatic Volume to report monthly results
    prog_compressed = pd.DataFrame(np.zeros((36, 36)))
    for i in range(0,36):
        start_day = 21*i
        end_day = 21*(i+1)
        prog_compressed.iloc[i,:] = prog_hedges_zeroes.iloc[start_day,:]
    
    #Compute Objective/Defensive Hedges
    print("starting objectives")
    parameters = pd.read_excel("Inputs.xlsx", sheet_name='Parameters ' +str(strat))
    #To hedge or not? #Programmatics set to 0 or 1000 so objectives don't interact with months marked for programmatics
    #parameters.loc[parameters.Category == "C", "Hedge Percentage"] = 0
    parameters.loc[parameters.Category == "P", "Hedge Percentage"] = 0 
    #parameters.loc[parameters.Category == "C", "Obj Boundary"] = 0
    #parameters.loc[parameters.Category == "C", "Defensive Boundary"] = 1000
    parameters.loc[parameters.Category == "P", "Obj Boundary"] = 0
    parameters.loc[parameters.Category == "P", "Defensive Boundary"] = 1000
    boundaries_compute = parameters.iloc[:,:]
    
    #Read in boundaries from excel file
    obj_bound_strat = pd.read_excel("Boundaries_Calc.xlsx", sheet_name='S' +str(strat)+' Obj Input')
    obj_bound_strat.columns = range(obj_bound_strat.shape[1])
    obj_bound_strat = obj_bound_strat.reset_index(drop = True)
    obj_bound_strat_final = obj_bound_strat.copy(deep = True)
    obj_bound_strat_final = obj_bound_strat_final.iloc[:,1:len(obj_bound_strat_final.columns)]
    obj_bound_strat_final.columns = range(obj_bound_strat_final.shape[1])
    dfen_bound_strat = pd.read_excel("Boundaries_Calc.xlsx", sheet_name='S' +str(strat)+' Dfen Input')
    dfen_bound_strat.columns = range(dfen_bound_strat.shape[1])
    dfen_bound_strat = dfen_bound_strat.reset_index(drop = True)
    dfen_bound_strat_final = dfen_bound_strat.copy(deep = True)
    dfen_bound_strat_final = dfen_bound_strat_final.iloc[:,1:len(dfen_bound_strat_final.columns)]
    dfen_bound_strat_final.columns = range(dfen_bound_strat_final.shape[1])

    
    obj_bounds = boundaries_compute.iloc[:,1].to_frame()
    obj_bounds_T = obj_bounds.transpose()
    def_bounds = boundaries_compute.iloc[:,3].to_frame()
    def_bounds_T = def_bounds.transpose()
    obj_hedge_target  = boundaries_compute.iloc[:,2].to_frame()
    obj_hedge_target_T = obj_hedge_target.transpose()
    
    obj_bounds_df = obj_bounds_T.append([obj_bounds_T]*(len(hedge_ratio_existing)-1),ignore_index=True)
    def_bounds_df = def_bounds_T.append([def_bounds_T]*(len(hedge_ratio_existing)-1),ignore_index=True)
    obj_hedge_target_df = obj_hedge_target_T.append([obj_hedge_target_T]*(len(hedge_ratio_existing)-1),ignore_index=True)
    
    obj_bounds_final = pd.DataFrame(0, index=np.arange(len(hedge_ratio_existing)), columns=hedge_ratio_existing.columns)
    def_bounds_final = pd.DataFrame(0, index=np.arange(len(hedge_ratio_existing)), columns=hedge_ratio_existing.columns)
    obj_hedge_target_final = pd.DataFrame(0, index=np.arange(len(hedge_ratio_existing)), columns=hedge_ratio_existing.columns)
    
    for k in range(0,len(prog_input)):
        days_start = 21*k
        days_end = 21* (k+1)
        obj_bounds_df_target = obj_bounds_df.iloc[days_start:days_end,:]
        obj_bounds_df_target = obj_bounds_df_target.shift(periods = k, axis = "columns", fill_value = 0)
        obj_bounds_final.iloc[days_start:days_end,:] = obj_bounds_df_target
        def_bounds_df_target = def_bounds_df.iloc[days_start:days_end,:]
        def_bounds_df_target = def_bounds_df_target.shift(periods = k, axis = "columns", fill_value = 0)
        def_bounds_final.iloc[days_start:days_end,:] = def_bounds_df_target
        obj_hedge_target_df_int = obj_hedge_target_df.iloc[days_start:days_end,:]
        obj_hedge_target_df_int = obj_hedge_target_df.shift(periods = k, axis = "columns", fill_value = 0)
        obj_hedge_target_final.iloc[days_start:days_end,:] = obj_hedge_target_df_int
    
    obj_bounds_final = obj_bound_strat_final.copy(deep = True)
    def_bounds_final = dfen_bound_strat_final.copy(deep = True)
    
    #Get prices
    daily_price_alliters = daily_prices.iloc[1:,4:]
    daily_price_alliters = daily_price_alliters.reset_index(drop = True)
    
    iter = iterations
    obj_prog_cost = pd.Series([0]*iter)
    obj_cost_sum = pd.DataFrame(np.zeros((iter, 36)))
    obj_vol_sum = pd.DataFrame(np.zeros((iter, 36)))
    forced_cost_sum = pd.DataFrame(np.zeros((iter, 36)))
    forced_vol_sum = pd.DataFrame(np.zeros((iter, 36)))
    prog_hedges_summarized=pd.DataFrame()
    prog_hedges_summarized_final = pd.DataFrame()
    prog_prices_summarized=pd.DataFrame()
    prog_prices_summarized_final = pd.DataFrame()
    obj_hedges_summarized=pd.DataFrame()
    obj_hedges_summarized_final = pd.DataFrame()
    obj_prices_summarized=pd.DataFrame()
    obj_prices_summarized_final = pd.DataFrame()
    obj_lowerb_summarized=pd.DataFrame()
    obj_lowerb_summarized_final = pd.DataFrame()
    obj_upperb_summarized=pd.DataFrame()
    obj_upperb_summarized_final = pd.DataFrame()
    force_hedges_summarized=pd.DataFrame()
    force_hedges_summarized_final = pd.DataFrame()
    force_prices_summarized=pd.DataFrame()
    force_prices_summarized_final = pd.DataFrame()
    
    #parameterize forced
    parameters_f = pd.read_excel("Inputs.xlsx", sheet_name='Parameters ' +str(strat))
    index_obj = (parameters_f.iloc[:,4].values == "O").argmax()
    index_col = (parameters_f.iloc[:,4].values == "C").argmax()
    parameters_f.iloc[index_obj+1:,2] = 0
    parameter_forced = pd.DataFrame(np.zeros((len(parameters.iloc[:,2]), 1)))
    parameter_forced.iloc[index_col,0] =  parameters_f.iloc[index_col,2]
    parameter_forced.iloc[index_obj,0] =  parameters_f.iloc[index_obj,2]
    parm = parameter_forced.iloc[index_col:index_obj+1,0]
    adder = ((parm[index_col])-(parm[index_obj]))/(len(parm)-1)
    for k in range(0,len(parm)-1):
        parm[k+1]= parm[k] - adder
    
    for i in range(0,iter):
        iters_start = 756*i
        iters_end = 756*(i+1)
        daily_prices_iter = daily_price_alliters.iloc[iters_start:iters_end,:]
        daily_prices_iter=daily_prices_iter.reset_index(drop = True)
        daily_prices_iter.columns = obj_bounds_final.columns
        daily_prices_iter_unique_first = daily_prices_iter.drop_duplicates()
        daily_prices_iter_unique_last = daily_prices_iter.drop_duplicates(keep = 'last')
        obj_bool = daily_prices_iter.lt(obj_bounds_final) | daily_prices_iter.gt(def_bounds_final)
        obj_bool = obj_bool.astype(int)
        bool_fix = daily_prices_iter.copy(deep = True)
        bool_fix[bool_fix>0] = 1
        obj_bool = obj_bool.multiply(bool_fix)
        obj_bool_month_formatted = pd.DataFrame(0, index=np.arange(len(hedge_ratio_existing)), columns=hedge_ratio_existing.columns)
        obj_bool_vol_month_formatted = pd.DataFrame(0, index=np.arange(len(hedge_ratio_existing)), columns=hedge_ratio_existing.columns)
        #Loop through months
     
        for month in range (0,36):
            start_month = month*21
            end_month = (21*(month+1))
            obj_bool_month = obj_bool.iloc[start_month:end_month,:]
            obj_bool_month = obj_bool_month.reset_index(drop = True)
            obj_bool_month_formatted_monthly= obj_bool_month.iloc[:,:]
            obj_bool_month_formatted_monthly = pd.DataFrame(0, index=np.arange(len(obj_bool_month)), columns=obj_bool_month.columns)
            obj_bool_month_formatted_monthly = obj_bool_month_formatted_monthly.reset_index(drop = True)
            for j in range(0,36):
                obj_month = obj_bool_month.iloc[:,j]
                obj_month = obj_month.reset_index(drop = True)
                index = (obj_month.values == 1).argmax()
                obj_month_bool = obj_month.index>index
                obj_month_bool = obj_month_bool * 0        
                obj_month_bool[index] = 1
                obj_bool_month_formatted_monthly.iloc[:,j] = obj_month_bool
            temp = obj_bool_month_formatted.iloc[start_month:end_month,:]
            obj_bool_month_formatted_monthly.index = temp.index
            obj_bool_month_formatted.iloc[start_month:end_month,:] = obj_bool_month_formatted_monthly
    
        #Loop through days
        obj_hedge_fill_ratio = obj_hedge_target_final.sub(ratio_track)
        obj_hedge_fill_ratio[obj_hedge_fill_ratio < 0] = 0
        
        #Make obj_hedge_fill_ratio non incremental
        obj_hedge_fill_ratio_process = pd.DataFrame(0, index=np.arange(len(obj_hedge_fill_ratio)), columns=obj_hedge_fill_ratio.columns)
        for k in range(1,len(obj_hedge_fill_ratio)):
            obj_hedge_fill_ratio_process.iloc[k,:] = obj_hedge_fill_ratio.iloc[k,:] - obj_hedge_fill_ratio.iloc[k-1,:]
        
        obj_hedge_fill_ratio_process.iloc[0,:] = obj_hedge_fill_ratio.iloc[0,:]
        obj_hedge_fill_ratio_process[obj_hedge_fill_ratio_process < 0] = 0
        #obj_hedge_fill_ratio = obj_hedge_fill_ratio_process.copy(deep = True)
        
        obj_decision = obj_bool_month_formatted.multiply(obj_hedge_fill_ratio)
        obj_decision_vol = obj_decision.multiply(budget_volumes_table)
        obj_decision_daily = obj_bool.multiply(obj_hedge_fill_ratio)
        obj_decision_vol_daily = obj_decision_daily.multiply(budget_volumes_table)
        
        #Make obj_decision incremental
        obj_decision_copy = obj_decision_daily.copy(deep = True)
        obj_decision_copy2 = obj_decision_daily.copy(deep = True)
        
        for x in range(0,len(obj_decision_copy.columns)):
            series_process = obj_decision_copy.iloc[:,x]
            series_process_uniq = series_process.drop_duplicates(keep = 'first')
            series_process_uniq = series_process_uniq.sort_values(ascending=True)
            roll_diff = pd.Series(series_process_uniq).diff(periods=1)
            series_process = series_process - series_process
            series_process = series_process.combine(roll_diff,max)
            obj_decision_copy2.iloc[:,x] = series_process
        
        obj_decision_copy2.iloc[0,:] = obj_decision_copy.iloc[0,:] #Output
        
        obj_decision_vol_bool = obj_decision_copy2.divide(obj_decision_copy2)
        obj_decision_vol_bool = obj_decision_vol_bool.fillna(0)
        obj_cost_prices = obj_decision_vol_bool.multiply(daily_prices_iter) 
        obj_cost_prices = obj_decision_vol_bool.multiply(daily_prices_iter) 
        obj_decision_copy2_vol = obj_decision_copy2.multiply(budget_volumes_table)
        obj_cost_daily = obj_decision_copy2_vol.multiply(daily_prices_iter)#Output
        
        #Process for monthly adjustments
        obj_decision_copy2_bool = obj_decision_copy2.copy(deep = True)
        obj_decision_copy2_bool[obj_decision_copy2_bool>0] = 1
        obj_hedge_fill_ratio_dec = obj_hedge_fill_ratio.multiply(obj_decision_copy2_bool)
            #Compare obj_hedge_fill_ratio_dec and 
        obj_daily_decisions = obj_decision_copy2.multiply(obj_hedge_fill_ratio)
        
        for k in range(0, len(obj_hedge_fill_ratio_dec.columns)):
            series = obj_hedge_fill_ratio_dec.iloc[:,k]
            series_list = series.tolist()
            if np.count_nonzero(series) == 0:
                series = series
            else:
                last_index = np.max(np.nonzero(series_list))
                first_index = np.min(np.nonzero(series_list))
                series[last_index:] = series_list[last_index]
                series[first_index:last_index] = series_list[first_index]
            obj_hedge_fill_ratio_dec.iloc[:,k] = series
        
        monthly_h2 = obj_hedge_fill_ratio_dec.copy(deep = True)
        for k in range(0,len(obj_hedge_fill_ratio_dec.columns)):
            start_day = 21*k
            end_day = 21*(k+1)-1
            monthly_h = obj_hedge_fill_ratio_dec.iloc[start_day:end_day,:]
            for x in range(0,len(monthly_h)):
                monthly_series = monthly_h.iloc[:,x]
                max_val = monthly_series.max()
                monthly_series[monthly_series>-1] = max_val
                monthly_h2.iloc[start_day:end_day,x] = monthly_series
        
        obj_hedge_fill_ratio_bool = obj_hedge_fill_ratio.copy(deep = True)
        obj_hedge_fill_ratio_bool[obj_hedge_fill_ratio_bool>0] = 1
        monthly_h2 = monthly_h2.multiply(obj_hedge_fill_ratio_bool)
        obj_decision = obj_bool_month_formatted.multiply(obj_hedge_fill_ratio_bool)
        obj_decision = obj_decision.multiply(monthly_h2)
        obj_decision_vol = obj_decision.multiply(budget_volumes_table)
        
        #Start processing to get incremental volumes for obj
        obj_compressed_vols = pd.DataFrame(np.zeros((36, 36)))
        obj_compressed_prices = pd.DataFrame(np.zeros((36, 36)))
        for k in range(0,36):
            start_day = 21*k
            end_day = 21*(k+1)-1
            obj_vol_month = obj_decision_vol.iloc[start_day:end_day,:]
            obj_vol_month_single = obj_vol_month.sum(axis = 0)
            obj_vol_month_single_T = obj_vol_month_single.transpose()
            obj_compressed_vols.iloc[k,:] = obj_vol_month_single
            obj_prices_month = obj_cost_prices.iloc[start_day:end_day,:]
            obj_prices_month_single = obj_prices_month.sum(axis = 0)
            obj_prices_month_single_T = obj_prices_month_single.transpose()
            obj_compressed_prices.iloc[k,:] = obj_prices_month_single
            
        obj_compressed_vols_incremental = pd.DataFrame(np.zeros((36, 36)))
        for k in range(0,len(obj_compressed_vols)-1):
            obj_compressed_vols_incremental.iloc[k,:] = obj_compressed_vols.iloc[k+1,:]-obj_compressed_vols.iloc[k,:]
        obj_compressed_vols_incremental[obj_compressed_vols_incremental<0] = 0
        obj_compressed_vols_incremental = obj_compressed_vols_incremental.shift(periods = 1)
        obj_compressed_vols_incremental.iloc[0,:] = obj_compressed_vols.iloc[0,:]
        
        obj_cost = obj_compressed_prices.multiply(obj_compressed_vols_incremental) #Objective hedge cost
        #Create Monthly Cost
        for k in range(0,36):
            obj_cost_monthly = obj_cost_daily.iloc[:,k]
            obj_cost_sum.iloc[i,k] = obj_cost_monthly.values.sum()
            obj_vol_monthly = obj_decision_copy2_vol.iloc[:,k]
            obj_vol_sum.iloc[i,k] = obj_vol_monthly.values.sum()
            
    #Start shift
        #Consolidate volume
        hedge_track_obj = obj_decision_vol.copy(deep = True)
        #hedge_track_obj = hedge_track.add(obj_decision_vol)
        hedge_track_obj_adjusted = pd.DataFrame(0, index=np.arange(len(month_diff)), columns=month_diff.columns)
        for k in range(0,36):
            start_day = 21*k
            end_day = 21*(k+1)-1
            hedge_track_month = hedge_track_obj.iloc[start_day:end_day+1,:]
            hedge_track_month = hedge_track_month.reset_index(drop = True)
            hedge_track_month_last = hedge_track_month.max(axis = 0)
            hedge_track_month_last_T = hedge_track_month_last.transpose()
            hedge_track_month_last_matrix = hedge_track_month_last_T.append([hedge_track_month_last_T]*(end_day - start_day),ignore_index=True)
            for j in range(start_day,end_day+1):
                hedge_track_obj_adjusted.iloc[j,:] = hedge_track_month_last
        
        #Incremental adjusted obj vol
        hedge_df = pd.DataFrame(0, index=np.arange(len(hedge_track_obj_adjusted)), columns=hedge_track_obj_adjusted.columns)
        hedge_series = pd.DataFrame(np.zeros((1,len(hedge_track_obj_adjusted))))
        hedge_series_temp = []
        hedge_track_obj_adjusted_copy = hedge_track_obj_adjusted.copy(deep = True)
        
        for x in range(0,len(hedge_track_obj_adjusted_copy.columns)):
            hedge_series_temp= hedge_track_obj_adjusted_copy.iloc[:,x]
            for y in range(1,len(hedge_series_temp)):
                hedge_series[y] = hedge_series_temp[y] - hedge_series_temp[y-1]
            hedge_series = hedge_series.transpose()
            hedge_df.iloc[:,x] =  hedge_series
        
        hedge_df[hedge_df<0] = 0    
        hedge_df.iloc[0,:] = hedge_track_obj_adjusted.iloc[0,:]
        
        hedge_df1 = pd.DataFrame(0, index=np.arange(len(hedge_track_obj_adjusted)), columns=hedge_track_obj_adjusted.columns)
        hedge_series1 = pd.DataFrame(np.zeros((len(hedge_track_obj_adjusted), 1)))
        
        for x in range(0,len(hedge_track_obj_adjusted.columns)):
            hedge_series_temp1 = hedge_df.iloc[:,x]
            for y in range(1,len(hedge_series_temp1)):
                hedge_series_temp1[y] = hedge_series_temp1[y] + hedge_series_temp1[y-1]
            hedge_df1.iloc[:,x] =  hedge_series_temp1
        
        hedge_track_obj_adjusted2 = hedge_track.add(hedge_df1)
        ratio_track_obj = hedge_track_obj_adjusted2.divide(budget_volumes_table)
    #End shift
    
    #Obj Force 13th month
        # parameters_f = pd.read_excel("Inputs.xlsx", sheet_name='Parameters 1')
        # index_obj = (parameters_f.iloc[:,4].values == "O").argmax()
        # index_col = (parameters_f.iloc[:,4].values == "C").argmax()
        # parameters_f.iloc[index_obj+1:,2] = 0
        # parameter_forced = pd.DataFrame(np.zeros((len(parameters.iloc[:,2]), 1)))
        # parameter_forced.iloc[index_col,0] =  parameters_f.iloc[index_col,2]
        # parameter_forced.iloc[index_obj,0] =  parameters_f.iloc[index_obj,2]
        # parm = parameter_forced.iloc[index_col:index_obj+1,0]
        # adder = ((parm[index_col])-(parm[index_obj]))/(len(parm)-1)
        # for k in range(0,len(parm)-1):
        # parm[k+1]= parm[k] - adder
        
        forced_target_ratio = pd.DataFrame(0, index=np.arange(len(month_diff)), columns=month_diff.columns)
        
        for k in range(0,36):
            start_day = 21*k
            end_day = (21*(k+1))-1
            forced_target_ratio.iloc[end_day,:] = parm
        
        forced_target_ratio = forced_target_ratio.fillna(0)
        forced_target_final_ratios = pd.DataFrame(0, index=np.arange(len(hedge_ratio_existing)), columns=hedge_ratio_existing.columns)
        
        for k in range(0,36):
            days_start = 21*k
            days_end = 21* (k+1)
            forced_month_target = forced_target_ratio.iloc[days_start:days_end,:]
            forced_month_target = forced_month_target.shift(periods = k, axis = "columns", fill_value = 0)
            forced_target_final_ratios.iloc[days_start:days_end,:] = forced_month_target
            forced_month_target = []
            
        forced_prices_bool = forced_target_final_ratios.divide(forced_target_final_ratios)
        forced_prices_bool = forced_prices_bool.fillna(0)
        ratio_track_collar = ratio_track_obj.multiply(forced_prices_bool)
        forced_incremental_ratio = forced_target_final_ratios.subtract(ratio_track_collar)
        forced_incremental_ratio[forced_incremental_ratio<0] = 0
        forced_incremental_vol = forced_incremental_ratio.multiply(budget_volumes_table)
        
        #Keep last value of each column from ratio track collar
        row_ix = ratio_track_collar.shape[0]-ratio_track_collar.ne(0).values[::-1].argmax(0)-1
        first_max = ratio_track_collar.values[row_ix, range(ratio_track_collar.shape[1])]
        ratio_max = pd.DataFrame([first_max], columns=ratio_track_collar.columns)
        
        row_ix = forced_target_final_ratios.shape[0]-forced_target_final_ratios.ne(0).values[::-1].argmax(0)-1
        first_max = forced_target_final_ratios.values[row_ix, range(forced_target_final_ratios.shape[1])]
        forced_ratio_max = pd.DataFrame([first_max], columns=forced_target_final_ratios.columns)
        
        forced_incremental_ratio = forced_ratio_max.subtract(ratio_max)
        forced_incremental_volume = forced_incremental_ratio.multiply(budget_volumes_table.iloc[0,:])
        forced_incremental_volume = forced_incremental_volume.round().abs()
        
        forced_incremental_volume_matrix = pd.DataFrame(0, index=np.arange(len(forced_target_final_ratios)), columns=forced_target_final_ratios.columns)
        
        for x in range(0,len(forced_incremental_volume.columns)):
            series = ratio_track_collar.iloc[:,x]
            series_list = series.tolist()
            row_id = np.max(np.nonzero(series_list))
            zeroes = np.zeros(len(series))
            zeroes[row_id] = forced_incremental_volume.iloc[0,x]
            forced_incremental_volume_matrix.iloc[:,x] = zeroes
        
        forced_incremental_vol = forced_incremental_volume_matrix.copy(deep = True)
        
        #Consolidate volumes
        forced_compressed = pd.DataFrame(np.zeros((36, 36)))
        for k in range(0,35):
            start_day = 21*k
            end_day = 21*(k+1)-1
            forced_compressed.iloc[k,:] = forced_incremental_vol.iloc[end_day,:]
            
        #Cost for forced
        forced_prices_zeroes = pd.DataFrame(0, index=np.arange(len(hedge_ratio_existing)), columns=hedge_ratio_existing.columns)
        daily_price_alliters_forced = daily_prices.iloc[1:,4:]
        daily_price_alliters_forced = daily_price_alliters_forced.reset_index(drop = True)
        #iter = 1
        #forced_cost_sum = pd.DataFrame([0])
        #forced_cost_sum = forced_cost_sum.transpose()
        #for i in range(0,iter):
        #iters_start = 756*i
        #iters_end = 756*(i+1)
        #daily_prices_iter = daily_price_alliters_forced.iloc[iters_start:iters_end,:]
        #daily_prices_iter.columns = forced_prices_zeroes.columns
        #Loop over days
        for j in range(0,34):
            day_start = 22*j
            day_end = 22* (j+1)
            forced_prices_zeroes.iloc[day_start,:] = daily_prices_iter.iloc[day_start,:]
        
        forced_cost = daily_prices_iter.multiply(forced_incremental_vol)
        
        for k in range(0,36):
            forced_cost_monthly = forced_cost.iloc[:,k]
            forced_cost_sum.iloc[i,k] = forced_cost_monthly.values.sum()
            forced_vol_monthly = forced_incremental_vol.iloc[:,k]
            forced_vol_sum.iloc[i,k] = forced_vol_monthly.values.sum()
    
        #Make compressed vols cumulative
        forced_series = pd.DataFrame(0, index=np.arange(len(forced_compressed)), columns=forced_compressed.columns)
        obj_series = pd.DataFrame(0, index=np.arange(len(obj_compressed_vols_incremental)), columns=obj_compressed_vols_incremental.columns)
        prog_series = pd.DataFrame(0, index=np.arange(len(prog_compressed)), columns=prog_compressed.columns)
        for j in range(0, len(forced_compressed.columns)):
            forced_series_process = forced_compressed.iloc[:,j]
            obj_series_process = obj_compressed_vols_incremental.iloc[:,j]
            prog_series_process = prog_compressed.iloc[:,j]
            for k in range(1,len(forced_series_process)):
                forced_series_process[k] = forced_series_process[k]+forced_series_process[k-1]
                obj_series_process[k] = obj_series_process[k]+obj_series_process[k-1]
                prog_series_process[k] = prog_series_process[k]+prog_series_process[k-1]
            forced_series.iloc[:,j]=forced_series_process
            obj_series.iloc[:,j]=obj_series_process
            prog_series.iloc[:,j]=prog_series_process
        
        final_hedge_vol = existing_compressed + forced_series + obj_series + prog_series
        final_ratio_compressed = final_hedge_vol.divide(budget_compressed)
        
        #Process summary from dailies
        delivery_months = ["9/1/2020",	"10/1/2020",	"11/1/2020",	"12/1/2020",	"1/1/2021",	"2/1/2021",	"3/1/2021",	"4/1/2021",	"5/1/2021",	"6/1/2021",	"7/1/2021",	"8/1/2021",	"9/1/2021",	"10/1/2021",	"11/1/2021",	"12/1/2021",	"1/1/2022",	"2/1/2022",	"3/1/2022",	"4/1/2022",	"5/1/2022",	"6/1/2022",	"7/1/2022",	"8/1/2022",	"9/1/2022",	"10/1/2022",	"11/1/2022",	"12/1/2022",	"1/1/2023",	"2/1/2023",	"3/1/2023",	"4/1/2023",	"5/1/2023",	"6/1/2023",	"7/1/2023",	"8/1/2023"]
        prog_hedges_summary = prog_hedges_zeroes.copy(deep = True)
        prog_hedges_summary.columns = delivery_months
        prog_hedges_summary['Transaction Day'] = prog_hedges_summary.index
        prog_hedges_summary['Iteration'] = i
        prog_hedges_summarized= prog_hedges_summary.melt(id_vars=['Transaction Day','Iteration'], value_vars=delivery_months)
        prog_hedges_summarized_final = prog_hedges_summarized_final.append(prog_hedges_summarized)
        
        prog_prices_summary = daily_prices_iter.multiply((prog_hedges_zeroes.copy(deep = True))/(prog_hedges_zeroes.copy(deep = True)))
        prog_prices_summary.columns = delivery_months
        prog_prices_summary['Transaction Day'] = prog_prices_summary.index
        prog_prices_summary['Iteration'] = i
        prog_prices_summarized= prog_prices_summary.melt(id_vars=['Transaction Day','Iteration'], value_vars=delivery_months)
        prog_prices_summarized_final = prog_prices_summarized_final.append(prog_prices_summarized)
    
        obj_hedges_summary = obj_decision_copy2_vol.copy(deep = True)
        obj_hedges_summary.columns = delivery_months
        obj_hedges_summary['Transaction Day'] = obj_hedges_summary.index
        obj_hedges_summary['Iteration'] = i
        obj_hedges_summarized= obj_hedges_summary.melt(id_vars=['Transaction Day','Iteration'], value_vars=delivery_months)
        obj_hedges_summarized_final = obj_hedges_summarized_final.append(obj_hedges_summarized)
        
        obj_prices_summary = daily_prices_iter.multiply((obj_decision_copy2_vol.copy(deep = True))/(obj_decision_copy2_vol.copy(deep = True)))
        obj_prices_summary.columns = delivery_months
        obj_prices_summary['Transaction Day'] = obj_prices_summary.index
        obj_prices_summary['Iteration'] = i
        obj_prices_summarized= obj_prices_summary.melt(id_vars=['Transaction Day','Iteration'], value_vars=delivery_months)
        obj_prices_summarized_final = obj_prices_summarized_final.append(obj_prices_summarized)
        
        obj_lowerb_summary = obj_bounds_final.multiply((obj_decision_copy2_vol.copy(deep = True))/(obj_decision_copy2_vol.copy(deep = True)))
        obj_lowerb_summary.columns = delivery_months
        obj_lowerb_summary['Transaction Day'] = obj_lowerb_summary.index
        obj_lowerb_summary['Iteration'] = i
        obj_lowerb_summarized= obj_lowerb_summary.melt(id_vars=['Transaction Day','Iteration'], value_vars=delivery_months)
        obj_lowerb_summarized_final = obj_lowerb_summarized_final.append(obj_lowerb_summarized)
        
        obj_upperb_summary = def_bounds_final.multiply((obj_decision_copy2_vol.copy(deep = True))/(obj_decision_copy2_vol.copy(deep = True)))
        obj_upperb_summary.columns = delivery_months
        obj_upperb_summary['Transaction Day'] = obj_upperb_summary.index
        obj_upperb_summary['Iteration'] = i
        obj_upperb_summarized= obj_upperb_summary.melt(id_vars=['Transaction Day','Iteration'], value_vars=delivery_months)
        obj_upperb_summarized_final = obj_upperb_summarized_final.append(obj_upperb_summarized)
        
        force_hedges_summary = forced_incremental_vol.copy(deep = True)
        force_hedges_summary.columns = delivery_months
        force_hedges_summary['Transaction Day'] = force_hedges_summary.index
        force_hedges_summary['Iteration'] = i
        force_hedges_summarized= force_hedges_summary.melt(id_vars=['Transaction Day','Iteration'], value_vars=delivery_months)
        force_hedges_summarized_final = force_hedges_summarized_final.append(force_hedges_summarized)
        
        force_prices_summary = daily_prices_iter.multiply((forced_incremental_vol.copy(deep = True))/(forced_incremental_vol.copy(deep = True)))
        force_prices_summary.columns = delivery_months
        force_prices_summary['Transaction Day'] = force_prices_summary.index
        force_prices_summary['Iteration'] = i
        force_prices_summarized= force_prices_summary.melt(id_vars=['Transaction Day','Iteration'], value_vars=delivery_months)
        force_prices_summarized_final = force_prices_summarized_final.append(force_prices_summarized)
    
    #Iterative program cost
    program_cost = prog_cost_sum + obj_cost_sum +forced_cost_sum
    program_vol = prog_vol_sum + obj_vol_sum + forced_vol_sum
    
    #Process output details 
    #prog_hedges_summarized_final prog_prices_summarized_final obj_hedges_summarized_final obj_prices_summarized_final force_hedges_summarized_final force_prices_summarized_final
    prog_hedges_summarized_final['Prices'] = prog_prices_summarized_final['value']
    prog_hedges_summarized_final['Protocol'] = 'P'
    prog_hedges_summarized_final = prog_hedges_summarized_final.loc[~((prog_hedges_summarized_final['value'] == 0))]
    obj_hedges_summarized_final['Prices'] = obj_prices_summarized_final['value']
    obj_hedges_summarized_final['Lower Bound'] = obj_lowerb_summarized_final['value']
    obj_hedges_summarized_final['Upper Bound'] = obj_upperb_summarized_final['value']
    obj_hedges_summarized_final['Protocol'] = 'O'
    obj_hedges_summarized_final = obj_hedges_summarized_final.loc[~((obj_hedges_summarized_final['value'] == 0))]
    force_hedges_summarized_final['Prices'] = force_prices_summarized_final['value']
    force_hedges_summarized_final['Protocol'] = 'F'
    force_hedges_summarized_final = force_hedges_summarized_final.loc[~((force_hedges_summarized_final['value'] == 0))]
    
    prog_hedges_summarized_final = prog_hedges_summarized_final.rename(columns = {"value":"Volume"})
    obj_hedges_summarized_final = obj_hedges_summarized_final.rename(columns = {"value":"Volume"})
    force_hedges_summarized_final = force_hedges_summarized_final.rename(columns = {"value":"Volume"})
    all_hedges_summary = prog_hedges_summarized_final.append(obj_hedges_summarized_final)
    all_hedges_summary = all_hedges_summary.append(force_hedges_summarized_final)
    
    #Output Hedge instances
    book = load_workbook('Summary Output\Strategy '+str(strat)+'\Hedge Output.xlsx')
    writer = pd.ExcelWriter('Summary Output\Strategy '+str(strat)+'\Hedge Output.xlsx', engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    all_hedges_summary.to_excel(writer, "Hedge activity")
    writer.save()
    
    #Output summaries
    book = load_workbook('Summary Output\Strategy '+str(strat)+'\Compressed_volumes_summary.xlsx')
    writer = pd.ExcelWriter('Summary Output\Strategy '+str(strat)+'\Compressed_volumes_summary.xlsx', engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    existing_compressed.to_excel(writer, "Existing hedges")
    forced_series.to_excel(writer, "Forced hedges")
    obj_series.to_excel(writer, "Objective Hedges")
    prog_series.to_excel(writer, "Prog Hedges")
    budget_compressed.to_excel(writer, "Budget Vols")
    writer.save()
    
    #Output cost summaries
    book = load_workbook('Summary Output\Strategy '+str(strat)+'\Cost Summary.xlsx')
    writer = pd.ExcelWriter('Summary Output\Strategy '+str(strat)+'\Cost Summary.xlsx', engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    prog_cost_sum.to_excel(writer, 'Programmatic costs' +str(strat))
    obj_cost_sum.to_excel(writer, 'Objective costs'+str(strat))
    forced_cost_sum.to_excel(writer, 'Forced costs'+str(strat))
    program_cost.to_excel(writer, 'Program costs'+str(strat))
    writer.save()
    
    #Output volume summaries
    book = load_workbook('Summary Output\Strategy '+str(strat)+'\Vol Summary.xlsx')
    writer = pd.ExcelWriter('Summary Output\Strategy '+str(strat)+'\Vol Summary.xlsx', engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    prog_vol_sum.to_excel(writer, 'Programmatic Vol' +str(strat))
    obj_vol_sum.to_excel(writer, 'Objective Vol'+str(strat))
    forced_vol_sum.to_excel(writer, 'Forced Vol'+str(strat))
    program_vol.to_excel(writer, 'Program Vol'+str(strat))
    writer.save()
    
    # #Output dailies - all hourly level volume output for all years
    # book = load_workbook('Summary Output\Strategy '+str(strat)+'\Client Details Daily.xlsx')
    # writer = pd.ExcelWriter('Summary Output\Strategy '+str(strat)+'Client Details Daily.xlsx', engine='openpyxl') 
    # writer.book = book
    # writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    # existing_hedges_table.to_excel(writer, "Existing Hedges")
    
    # prog_hedges_zeroes.to_excel(writer, "Prog Hedges")
    # prog_cost.to_excel(writer, "Prog Cost")
    
    # obj_decision_copy2_vol.to_excel(writer, "Objective Hedges")
    # obj_cost_daily.to_excel(writer, "Obj Costs")
    
    # forced_cost.to_excel(writer, "Forced Hedges")
    # forced_incremental_vol.to_excel(writer, "Forced Hedge Cost")
    
    # daily_prices_iter.to_excel(writer, "Daily Prices")
    
    # writer.save()


# Process no hedge

open_volumes = (0.98*budget_volumes_table).subtract(existing_hedges_table)
open_vol_cost = pd.DataFrame(np.zeros((iterations, 36)))
open_vol_monthly = pd.DataFrame(np.zeros((iterations, 36)))

for i in range(0,iterations):
    iters_start = 756*i
    iters_end = 756*(i+1)
    daily_prices_iter = daily_price_alliters.iloc[iters_start:iters_end,:]
    daily_prices_iter.columns = open_volumes.columns
    daily_prices_iter.reset_index(inplace = True, drop = True)
    daily_prices_iter_first_month = daily_prices_iter.iloc[0:21,:]
    daily_prices_iter_first_month.columns = daily_prices_iter.columns
    daily_prices_iter_first_month = daily_prices_iter_first_month.reset_index(drop = True)
    
    #Value opoen at spot, comment out if value at day 1 fwds    
    for k in range(0,len(daily_prices_iter.columns)):
        daily_prices_series = daily_prices_iter.iloc[:,k]
        row_ix = daily_prices_series.shape[0]-daily_prices_series.ne(0).values[::-1].argmax(0)-1
        series = daily_prices_series[row_ix-20: row_ix+1]
        series = series.reset_index(drop=True)
        daily_prices_iter_first_month.iloc[:,k] = series
        
    open_volumes_rep = open_volumes.iloc[0:21,:]
    open_volumes_rep_mean = open_volumes_rep.mean()
    open_volumes_cost = open_volumes_rep.multiply(daily_prices_iter_first_month)
    open_volumes_cost_mean = open_volumes_cost.mean()
    open_vol_cost.iloc[i,:] = open_volumes_cost_mean
    open_vol_monthly.iloc[i,:] = open_volumes_rep_mean


book = load_workbook('Summary Output\Strategy '+str(strat+1)+'\Cost Summary.xlsx')
writer = pd.ExcelWriter('Summary Output\Strategy '+str(strat+1)+'\Cost Summary.xlsx', engine='openpyxl') 
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
open_vol_cost.to_excel(writer, 'Open Cost')
writer.save()

#Output volume summaries
book = load_workbook('Summary Output\Strategy '+str(strat+1)+'\Vol Summary.xlsx')
writer = pd.ExcelWriter('Summary Output\Strategy '+str(strat+1)+'\Vol Summary.xlsx', engine='openpyxl') 
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
open_vol_monthly.to_excel(writer, 'Open Vol')
writer.save()

elapsed = time.time() - t